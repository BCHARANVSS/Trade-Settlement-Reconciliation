import pandas as pd
import numpy as np
from openpyxl import load_workbook  
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  
from openpyxl.utils import get_column_letter  

# load data
df_trades = pd.read_csv("trade_data.csv")
df_break = pd.read_csv("breaks_identified.csv")

df_trades.columns = df_trades.columns.str.strip().str.lower().str.replace(" ","_")
df_break.columns = df_break.columns.str.strip().str.lower().str.replace(" ","_")

# prepare all sheets data 

# sheet 1 - executive summary

total_trades = len(df_trades)
total_break = len(df_break)
clean_trades = total_trades-total_break
break_rate = round((total_break/total_trades)*100,2)
total_cash_at_risk = round(df_break["abs_break_amount"].sum(),2)
high_risk_count = len(df_break[df_break["risk_level"]=="high"])
medium_risk_count = len(df_break[df_break["risk_level"]=="medium"])
low_risk_count = len(df_break[df_break["risk_level"]=="low"])
largest_break = round(df_break["abs_break_amount"].max(),2)
underpayments = len(df_break[df_break["break_direction"]=="underpayment"])
overpayment = len(df_break[df_break["break_direction"]=="overpayment"])

summary_data = {
    "metric":[
        "total trades",
        "clean trades",
        "total breaks found",
        "break risk (%)",
        "total cash at risk",
        "high risk breaks",
        "medium risk breaks",
        "low risk breaks",
        "largest single break",
        "total underpayment",
        "total over payment"
    ],
    "value":[
        total_trades,
        clean_trades,
        total_break,
        break_rate,
        total_cash_at_risk,
        high_risk_count,
        medium_risk_count,
        low_risk_count,
        largest_break,
        underpayments,
        overpayment
    ]
}
df_summary = pd.DataFrame(summary_data)

# sheet 2

df_escalations = df_break.nlargest(20,"abs_break_amount")[[
    "trade_id","counterparty","asset","trade_type",
    "currency","expected_amounts","settled_amount",
    "abs_break_amount","risk_level","break_direction"
]].reset_index(drop=True)

# sheet 3 counterparty analysis
df_counterparty = df_break.groupby("counterparty").agg(
    total_break=("trade_id","count"),
    total_cash_at_risk=("abs_break_amount","sum"),
    avg_break_amount = ("abs_break_amount","mean"),
    high_risk_break = ("risk_level",lambda x:(x== "high").sum()),
    underpayments = ("break_direction",lambda x:(x=="under payment").sum()),
    overpayment= ("break_direction",lambda x:(x=="over payment").sum())
).round(2).sort_values("total_cash_at_risk",ascending=False).reset_index()

# sheet 4 buy vs sell break analysis

df_buysell = df_break.groupby(["trade_type","break_direction"]).agg(
    total_break=("trade_id","count"),
    total_cash_at_risk=("abs_break_amount","sum"),
    avg_break = ("abs_break_amount","mean")
).round(2).reset_index()

# sheet 5 currency risk

df_currency = df_break.groupby("currency").agg(
    total_break=("trade_id","count"),
    total_cash_at_risk=("abs_break_amount","sum"),
    avg_break = ("abs_break_amount","mean")
).round(2).sort_values("total_cash_at_risk",ascending=False).reset_index()

# sheet 6 full break resgister

df_full = df_break[[
    "trade_id","counterparty","asset","trade_type",
    "currency","expected_amounts","settled_amount",
    "break_amount","abs_break_amount",
    "risk_level","break_direction"
]].sort_values("abs_break_amount",ascending=False).reset_index(drop=True)


# write to excel 

output_file = "GS_break_report.xlsx"

with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:


    df_summary.to_excel(writer,sheet_name="executive summary",index=False)
    df_escalations.to_excel(writer,sheet_name="top escalation",index=False)
    df_counterparty.to_excel(writer,sheet_name="counterparty analysis",index=False)
    df_buysell.to_excel(writer,sheet_name="buy vs sell analysis",index=False)
    df_currency.to_excel(writer,sheet_name="currency risk",index=False)
    df_full.to_excel(writer,sheet_name="full break register",index=False)

    workbook = writer.book

    ## formats ##

    # header format - dark blue background, white bold text

    header_fmt = workbook.add_format({
        "bold": True,
        "font_color": "white",
        "bg_color": "#003366",
        "align" : "center",
        "valign" : "vcenter",
        "border":1
    })

    # number format comma seperated,2 decimal places
    number_fmt = workbook.add_format({
        "num_format":"#,##0.00",
        "border":1
    })

    # text format 
    text_fmt = workbook.add_format({
        "border":1
    })

    # high risk - red background

    high_fmt = workbook.add_format({
        "bg_color": "#FF4444",
        "font_color": "white",
        'bold': True,
        "border" : 1
    })

    # medium risk - orange background

    medium_fmt = workbook.add_format({
        "bg_color" : "#FFA500",
        "font_color": "white",
        "bold" : True,
        'border': 1
    })

    # low risk -  green background
    low_fmt = workbook.add_format({
        "bg_color" : "#00AA00",
        "font_color" : "white",
        "bold" : True,
        "border" : 1
    })


### format of each sheet ###

# executive summary sheet

ws_summary = writer.sheets["executive summary"]
ws_summary.set_column("A:A",30)
ws_summary.set_column("B:B",25)
ws_summary.write(0,0,"metric",header_fmt)
ws_summary.write(0,1,"value",header_fmt)

for row_num in range(1,len(df_summary)+1):
    ws_summary.write(row_num,0,df_summary.iloc[row_num-1]["metric"],text_fmt)
    ws_summary.write(row_num,1,df_summary.iloc[row_num-1]["value"],number_fmt)

# top escalation sheet

ws_esc = writer.sheets["top escalation"]
for col_num,col_name in enumerate(df_escalations.columns):
    ws_esc.write(0,col_num,col_name,header_fmt)
    ws_esc.set_column(col_num,col_num,20)
for row_num in range(len(df_escalations)):
    risk = df_escalations.iloc[row_num]["risk_level"]
    risk_fmt = high_fmt if risk == "high" else medium_fmt if risk == "medium" else low_fmt
    for col_num in range(len(df_escalations.columns)):
        val = df_escalations.iloc[row_num,col_num]
        if isinstance(val,(int,float)):
            ws_esc.write(row_num+1,col_num, val,number_fmt)
        else:
            ws_esc.write(row_num+1,col_num, val,text_fmt)
    ws_esc.write(
        row_num+1,
        df_escalations.columns.get_loc("risk_level"),
        risk,
        risk_fmt
    )

# counterparty analysis sheet
ws_cp = writer.sheets["counterparty analysis"]
for col_num, col_name in enumerate(df_counterparty.columns):
    ws_cp.write(0, col_num, col_name, header_fmt)
    ws_cp.set_column(col_num, col_num, 22)

for row_num in range(len(df_counterparty)):
    for col_num in range(len(df_counterparty.columns)):
        val = df_counterparty.iloc[row_num, col_num]
        if isinstance(val, (int, float)):
            ws_cp.write(row_num + 1, col_num, val, number_fmt)
        else:
            ws_cp.write(row_num + 1, col_num, val, text_fmt)

# buy vs sell sheet 

ws_bs = writer.sheets["buy vs sell analysis"]
for col_num,col_name in enumerate(df_buysell.columns):
    ws_bs.write(0,col_num,col_name,header_fmt)
    ws_bs.set_column(col_num,col_num,22)

for row_num in range(len(df_buysell)):
    for col_num in range (len(df_buysell.columns)):
        val=df_buysell.iloc[row_num,col_num]
        if isinstance(val,(int,float)):
            ws_bs.write(row_num+1,col_num,val,number_fmt)
        else:
            ws_bs.write(row_num+1,col_num,val,text_fmt)

# currency risk sheet 

ws_curr = writer.sheets["currency risk"]
for col_num,col_name in enumerate(df_currency.columns):
    ws_curr.write(0,col_num,col_name,header_fmt)
    ws_curr.set_column(col_num,col_num,22)

for row_num in range(len(df_currency)):
    for col_num in range(len(df_currency.columns)):
        val = df_currency.iloc[row_num,col_num]
        if isinstance(val,(int,float)):
            ws_curr.write(row_num+1,col_num,val,number_fmt)
        else:
            ws_curr.write(row_num+1,col_num,val,text_fmt)

# full break register sheet 

ws_full = writer.sheets["full break register"]
for col_num,col_name in enumerate(df_full.columns):
    ws_full.write(0,col_num,col_name,header_fmt)
    ws_full.set_column(col_num,col_num,20)

for row_num in range(len(df_full)):
    for col_num in range(len(df_full.columns)):
        val=df_full.iloc[row_num,col_num]
        if isinstance(val,(int,float)):
            ws_full.write(row_num+1,col_num,val,text_fmt)
        else:
            ws_full.write(row_num+1,col_num,val,text_fmt)

print("="*50)
print("EXCEL REPORT GENERATED")
print("="*50)
print(f"file saved:{output_file}")
print(f"sheets created;")
print(f" 1.executive summary")
print(f" 2.top escalations - colour by risk")
print(f" 3.counterparty analysis")
print(f" 4.buy vs sell analysis")
print(f" 5.currency risk")
print(f" 6.full break register {len(df_full)}break")
print("="*50)